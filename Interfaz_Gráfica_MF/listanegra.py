#Importa librerias
from openpyxl import load_workbook
import os 
from tkinter import messagebox as MessageBox
#Funcion encargada de añadir nombres a la lista negra
def añadirlistanegra(nombre):   #La función recibe una cadena para trabajar
    ruta="Base de datos/"  #define la ruta de la carpeta que ocntiene los rostros
    
    archivo= load_workbook('listaexcluidos.xlsx')  #Abre el archivo y lo almacena en la variable archivo
    ws=archivo["Hoja1"]  #Guarda la hoja 1 del archivo excel en la variable
    cell_range=ws['A2':'A100']  #Define un rango de las celdas y lo almacena en una variable
    
#ayuda a buscar y comparar la cadena que le llegó, con la lista de nombres de la base de datos
#En el caso de que la cadena ingresada coincida con algun nombre de la base de datos
#se añade a la lista negra en un espacio en blanco en el excel
    for base in os.listdir(ruta):  
        
        if base == nombre:
            if buscarenlista(nombre)==True:
                break
            else:
                for i in cell_range:
                    
                    if (i[0].value == None):
                        
                        i[0].value=nombre
                        break
            archivo.save('listaexcluidos.xlsx') #Guarda lo realizado en el excel
            break
        
#Funcion encargada de eliminar nombres a la lista negra
def eliminardelista(nombre):
    archivo= load_workbook('listaexcluidos.xlsx')
    ws=archivo["Hoja1"]
    cell_range=ws['A2':'A100']
#ayuda a buscar la cadena recibida y comparar ocn los datos de excel
#si encunetra algun similitud, elimina el dato que es igual
    for i in cell_range:
        
        if (i[0].value == nombre):
            
            i[0].value=""
    archivo.save('listaexcluidos.xlsx')


#Funcion encargada de buscar nombres en la lista negra
def buscarenlista(nombre):
    archivo= load_workbook('listaexcluidos.xlsx')
    ws=archivo["Hoja1"]
    cell_range=ws['A2':'A100']
    contador=0

# Busca en el archivo excel el nombre recibido en la cadena, y si existe, devuelve verdadero
# si no existe devuelve falso
    for i in cell_range:
        
        contador=contador+1
        
        if (i[0].value == nombre):
            
            return True
        else:
            if contador==98:
                return False

